import os
from openpyxl import Workbook
from time import strftime,localtime,time
from openpyxl.chart import LineChart,Reference



def read_file():
    """
    读取dump下来的数据
    """
    with open('dump_meminfo.txt') as f:
        data_list = []
        while True:  
            line = f.readline()
            if not line:
                break
            content = line.split()  # 已空格划分，生成列表
            try:
                if len(content) == 12:
                    del content[0]  # 删除第一个元素
                else:
                    del content[:2]
                content.pop()  # 删除最后一个元素
            except IndexError:
                pass
            if content:
                data_list.append(content)
        return data_list


def add_line_chart(ws):
    c1=LineChart()                                        #实例化创建 折线图表实例对象
    c1.title='趋势图'                                 #图表标题
    c1.style=13                                           #图表样式
    c1.x_axis.title='次数'                         #图表X轴标题
    c1.y_axis.title='value'                                #图表Y轴标题
    data=Reference(ws, min_col=5, max_col=6, min_row=1, max_row=ws.max_row)  #引用数据系列，用作折线图表的数据值。
    c1.add_data(data, titles_from_data=True)              #图表添加数据系列。
    ws.add_chart(c1, 'L12')


def create_report():
    """
    生成excel表格
    """
    wb = Workbook()
    ws = wb.active
    rows = read_file()
    title = ['USER','PR','NI','VIRT','RES(M)','SHR(M)','S','%CPU','%MEM','TIME+']
    ws.append(title)
    for row in rows:
        row[4] = int(row[4][:-1])
        row[5] = int(row[5][:-1])
        row[7] = float(row[7])
        row[8] = float(row[8])
        ws.append(row)
    
    description = ['Description:',
              'PR 优先级',
              'NI nice值。负值表示高优先级，正值表示低优先级',
              'VIRT 进程使用的虚拟内存总量，单位kb。VIRT=SWAP+RES',
              'RES 进程使用的、未被换出的物理内存大小，单位kb。RES=CODE+DATA',
              'SHR 共享内存大小，单位kb',
              'S 进程状态（D=不可中断的睡眠状态，R=运行，S=睡眠，T=跟踪/停止，Z=僵尸进程）',
              '%CPU 上次更新到现在的CPU时间占用百分比',
              '%MEM 进程使用的物理内存百分比',
              'TIME+ 进程使用的CPU时间总计，单位1/100秒']

    for i in range(10):
        ws['L{}'.format(i+1)] = description[i] 

    add_line_chart(ws)
    
    file_name =strftime('%Y%m%d%H%M%S',localtime()) + '_report'
    wb.save('{}.xlsx'.format(file_name))


def main():
    create_report()


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(e)
